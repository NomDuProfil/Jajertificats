# -*- coding: utf-8 -*-

import socket, ssl, datetime, OpenSSL, sys, os, multiprocessing, re, configparser, glob, string
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import PieChart, Reference, label, Series
from time import sleep

DUREE_ENTRE_CHAQUE_REQUETE = 0 #En seconde
TIMEOUT_REQUEST = 5 #En seconde
EXPIRATION_DAYS = 30 #En nombre de jour
PERIODE_DE_VALIDITE = 825 #En nombre de jour
DEBUG = False #Mode DEBUG

LIST_PORT = ['443']

def checkconfigfile(path):
	config = configparser.ConfigParser()
	config.sections()
	config.read(path)
	for currentsection in config.sections():
		for current in config.items(currentsection):
			if config.get(currentsection, current[0]) == "":
				print '\033[1;31m[ERREUR FICHIER CONF]\033[1;m'+" Le parametre "+current[0]+" de la section "+currentsection+" est vide"
				sys.exit(1)

def search_start_line(worksheet):
	currentline = 1
	while not (worksheet["B"+str(currentline)].value == "Domaines" or worksheet["B"+str(currentline)].value == "Domain(s)"):
		currentline = currentline+1
	return currentline+1

def sortresult(result):
	statusarray = {"ok": [], "wildcard": [], "expire": [], "expiresoon": [], "expiresoon": [], "validitytoolong": [], "notmatch": [], "timeout": [], "errresolution": [], "error": []}
	for current in result:
		try:
			statusarray[current["status"]].append(current)
		except:
			statusarray["error"].append(current)
	finaltab = statusarray["expire"]+statusarray["expiresoon"]+statusarray["validitytoolong"]+statusarray["notmatch"]+statusarray["wildcard"]+statusarray["ok"]+statusarray["errresolution"]+statusarray["error"]
	return finaltab

def getstatuscolor(status):
	if status == "ok":
		return '18b000'
	elif status == "wildcard":
		return '18b000'
	elif status == "expire":
		return 'ff0000'
	elif status == "expiresoon":
		return 'dd6400'
	elif status == "validitytoolong":
		return 'fdff00'
	elif status == "notmatch":
		return '00b5ea'
	return 'ffffff'

def getstatustextcolor(status):
	if status == "ok":
		return 'ffffff'
	elif status == "wildcard":
		return 'ffffff'
	elif status == "expire":
		return 'ffffff'
	elif status == "expiresoon":
		return 'ffffff'
	elif status == "validitytoolong":
		return '000000'
	elif status == "notmatch":
		return '000000'
	return 'ffffff'

def getstatustext(status):
	if status == "ok":
		return "VALIDE"
	elif status == "wildcard":
		return "WILDCARD"
	elif status == "expire":
		return u"EXPIRÉS"
	elif status == "expiresoon":
		return u"EXPIRENT BIENTÔT"
	elif status == "validitytoolong":
		return "VALIDITE TROP LONGUE"
	elif status == "notmatch":
		return u"NOM D'HÔTE INVALIDE"
	return 'Status inconnu'

def writeresult(result, pathsave):
	result = sortresult(result)
	statuscount = {"ok": 0, "wildcard": 0, "expire": 0, "expiresoon": 0, "expiresoon": 0, "validitytoolong": 0, "notmatch": 0, "timeout": 0, "errresolution": 0, "error": 0}
	wb = Workbook()
	wssommaire = wb.active
	wssommaire.title = u'Sommaire'
	wssommaire.column_dimensions['A'].width = len("VALIDITE TROP LONGUE")*1.3
	wssommaire["A1"] = u"EXPIRÉ"
	wssommaire["A1"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("expire"), end_color=getstatuscolor("expire"))
	wssommaire["A1"].font = Font(color='FFFFFF')
	wssommaire["A2"] = u"EXPIRE BIENTÔT"
	wssommaire["A2"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("expiresoon"), end_color=getstatuscolor("expiresoon"))
	wssommaire["A2"].font = Font(color='FFFFFF')
	wssommaire["A3"] = "VALIDITE TROP LONGUE"
	wssommaire["A3"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("validitytoolong"), end_color=getstatuscolor("validitytoolong"))
	wssommaire["A3"].font = Font(color='000000')
	wssommaire["A4"] = u"NOM D'HÔTE INVALIDE"
	wssommaire["A4"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("notmatch"), end_color=getstatuscolor("notmatch"))
	wssommaire["A4"].font = Font(color='000000')
	wssommaire["A5"] = "WILDCARD"
	wssommaire["A5"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("wildcard"), end_color=getstatuscolor("wildcard"))
	wssommaire["A5"].font = Font(color='FFFFFF')
	wssommaire["A6"] = "VALIDES"
	wssommaire["A6"].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("ok"), end_color=getstatuscolor("ok"))
	wssommaire["A6"].font = Font(color='FFFFFF')
	ws = wb.create_sheet(u'Détails')
	listwidthcolumn = [] 
	ws['A1'] = "Status certificat"
	ws['A1'].alignment = Alignment(horizontal='center')
	ws['A1'].font = Font(bold=True)
	ws.column_dimensions['A'].width = len("Status certificat")*1.3
	ws['B1'] = "Domaines"
	ws['B1'].alignment = Alignment(horizontal='center')
	ws['B1'].font = Font(bold=True)
	ws['C1'] = "Port"
	ws.column_dimensions['C'].width = len("Port")*1.3
	ws['C1'].alignment = Alignment(horizontal='center')
	ws['C1'].font = Font(bold=True)
	ws['D1'] = "Date de délivrance"
	ws.column_dimensions['D'].width = len("Date de délivrance")*1.3
	ws['D1'].alignment = Alignment(horizontal='center')
	ws['D1'].font = Font(bold=True)
	ws['E1'] = "Date d'expiration"
	ws.column_dimensions['E'].width = len("Date d'expiration")*1.3
	ws['E1'].alignment = Alignment(horizontal='center')
	ws['E1'].font = Font(bold=True)
	ws['F1'] = "Jours restants"
	ws.column_dimensions['F'].width = len("Jours restants")*1.3
	ws['F1'].alignment = Alignment(horizontal='center')
	ws['F1'].font = Font(bold=True)
	ws['G1'] = "Validité (nombre de jours)"
	ws.column_dimensions['G'].width = len("Validité (nombre de jours)")*1.3
	ws['G1'].alignment = Alignment(horizontal='center')
	ws['G1'].font = Font(bold=True)
	ws['H1'] = "Vérifié par"
	ws.column_dimensions['H'].width = len("Emis par")*1.3
	ws['H1'].alignment = Alignment(horizontal='center')
	ws['H1'].font = Font(bold=True)
	ws['I1'] = "Emis pour"
	ws.column_dimensions['I'].width = len("Emis pour")*1.3
	ws['I1'].alignment = Alignment(horizontal='center')
	ws['I1'].font = Font(bold=True)
	ws['J1'] = "Numéro de série"
	ws.column_dimensions['J'].width = len("Numéro de série")*1.3
	ws['J1'].alignment = Alignment(horizontal='center')
	ws['J1'].font = Font(bold=True)
	currentline = 2
	wsnotmatch = wb.create_sheet(u'Noms d\'hôte invalides')
	wsnotmatch['A1'] = "Domaines"
	wsnotmatch['B1'] = "Port"
	wserrresolution = wb.create_sheet(u'Erreur de résolution')
	wserrresolution['A1'] = "Domaines"
	wserrresolution['B1'] = "Port"
	wstimeout = wb.create_sheet(u'Timeout')
	wstimeout['A1'] = "Domaines"
	wstimeout['B1'] = "Port"
	wsinvalidecert = wb.create_sheet(u'Certificats invalides')
	wsinvalidecert['A1'] = "Domaines"
	wsinvalidecert['B1'] = "Port"
	wserror = wb.create_sheet(u'Autres erreurs')
	wserror['A1'] = "Domaines"
	wserror['B1'] = "Port"
	countersheet = {'principal': 1, 'notmatch': 1, 'errresolution': 1, 'timeout': 1, 'error': 1}
	currentlinenotmatch = 1
	currentsheetinfo = {'object': None, 'line': 0}
	for current in result:
		if current["status"] == "ok":
			statuscount["ok"] += 1
			countersheet["principal"] += 1
			ws['A'+str(countersheet["principal"])] = "VALIDE"
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("ok"), end_color=getstatuscolor("ok"))
			currentsheetinfo['object'] = ws
			currentsheetinfo['line'] = countersheet["principal"]
		elif current["status"] == "wildcard":
			statuscount["wildcard"] += 1
			countersheet["principal"] += 1
			ws['A'+str(countersheet["principal"])] = "WILDCARD"
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("wildcard"), end_color=getstatuscolor("wildcard"))
			currentsheetinfo['object'] = ws
			currentsheetinfo['line'] = countersheet["principal"]
		elif current["status"] == "expire":
			statuscount["expire"] += 1
			countersheet["principal"] += 1
			ws['A'+str(countersheet["principal"])] = "EXPIRÉ"
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("expire"), end_color=getstatuscolor("expire"))
			currentsheetinfo['object'] = ws
			currentsheetinfo['line'] = countersheet["principal"]
		elif current["status"] == "expiresoon":
			statuscount["expiresoon"] += 1
			countersheet["principal"] += 1
			ws['A'+str(countersheet["principal"])] = "EXPIRE BIENTOT"
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("expiresoon"), end_color=getstatuscolor("expiresoon"))
			currentsheetinfo['object'] = ws
			currentsheetinfo['line'] = countersheet["principal"]
		elif current["status"] == "validitytoolong":
			statuscount["validitytoolong"] += 1
			countersheet["principal"] += 1
			if ws.column_dimensions['A'].width < len("VALIDITE TROP LONGUE")*1.3:
				ws.column_dimensions['A'].width = len("VALIDITE TROP LONGUE")*1.3
			ws['A'+str(countersheet["principal"])] = "VALIDITÉ TROP LONGUE"
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor("validitytoolong"), end_color=getstatuscolor("validitytoolong"))
			currentsheetinfo['object'] = ws
			currentsheetinfo['line'] = countersheet["principal"]
		elif current["status"] == "notmatch":
			statuscount["notmatch"] += 1
			countersheet["notmatch"] += 1
			if wsnotmatch.column_dimensions['A'].width < len("HOSTNAME INVALIDE")*1.3:
				wsnotmatch.column_dimensions['A'].width = len("HOSTNAME INVALIDE")*1.3
			currentsheetinfo['object'] = wsnotmatch
			currentsheetinfo['line'] = countersheet["notmatch"]
		else:
			if current["status"] == "errresolution":
				countersheet["errresolution"] += 1
				currentsheetinfo['object'] = wserrresolution
				currentsheetinfo['line'] = countersheet["errresolution"]
				statuscount["errresolution"] += 1
			elif current["status"] == "timeout":
				countersheet["timeout"] += 1
				statuscount["timeout"] += 1
				currentsheetinfo['line'] = countersheet["timeout"]
				currentsheetinfo['object'] = wstimeout
			else:
				countersheet["error"] += 1
				statuscount["error"] += 1
				currentsheetinfo['line'] = countersheet["error"]
				currentsheetinfo['object'] = wserror
		if currentsheetinfo['object'] == ws:
			ws['A'+str(countersheet["principal"])] = getstatustext(current["status"])
			ws['A'+str(countersheet["principal"])].fill = PatternFill(fill_type="solid", start_color=getstatuscolor(current["status"]), end_color=getstatuscolor(current["status"]))
			ws['A'+str(countersheet["principal"])].font = Font(color=getstatustextcolor(current["status"]))
			ws['A'+str(countersheet["principal"])].alignment = Alignment(horizontal='center')
			ws['B'+str(countersheet["principal"])] = current["domain"]
			if ws.column_dimensions['B'].width < len(current["domain"])*1.3 and ws.column_dimensions['B'].width != 255:
				if len(current["domain"])*1.3 >= 255:
					ws.column_dimensions['B'].width = 255
				else:
					ws.column_dimensions['B'].width = len(current["domain"])*1.3
			if current["status"] == "errresolution":
				ws['C'+str(countersheet["principal"])] = '*'
			else:
				ws['C'+str(countersheet["principal"])] = current["port"]
			if ws.column_dimensions['C'].width < len(current["port"])*1.3:
				ws.column_dimensions['C'].width = len(current["port"])*1.3
			if (str(current["notBefore"]) != "error"):
				ws['D'+str(countersheet["principal"])] = str(current["notBefore"].year)+'-'+"{:02d}".format(current["notBefore"].month)+'-'+"{:02d}".format(current["notBefore"].day)
			else:
				ws['D'+str(countersheet["principal"])] = "ERROR"
			if (str(current["notAfter"]) != "error"):
				ws['E'+str(countersheet["principal"])] = str(current["notAfter"].year)+'-'+"{:02d}".format(current["notAfter"].month)+'-'+"{:02d}".format(current["notAfter"].day)
			else:
				ws['E'+str(countersheet["principal"])] = "ERROR"
			if(str(current["serialNumber"]) != "error"):
				if ws.column_dimensions['J'].width < len(str(current["serialNumber"]))*1.3:
					ws.column_dimensions['J'].width = len(str(current["serialNumber"]))*1.3
				ws['J'+str(countersheet["principal"])] = str(current["serialNumber"])
			else:
				ws['J'+str(countersheet["principal"])] = "ERROR"
			ws['F'+str(countersheet["principal"])] = str(current["deltaToday"])
			ws['G'+str(countersheet["principal"])] = str(current["periodevalidity"])
			ws['H'+str(countersheet["principal"])] = current["deliver"]
			if ws.column_dimensions['H'].width < len(current["deliver"])*1.3:
				ws.column_dimensions['H'].width = len(current["deliver"])*1.3
			ws['I'+str(countersheet["principal"])] = current["deliverfor"]
			if ws.column_dimensions['I'].width < len(current["deliverfor"])*1.3:
				ws.column_dimensions['I'].width = len(current["deliverfor"])*1.3
		else:
			currentsheetinfo['object']['A'+str(currentsheetinfo['line'])] = current["domain"]
			currentsheetinfo['object']['B'+str(currentsheetinfo['line'])] = current["port"]
	wssommaire["B1"] = statuscount["expire"]
	wssommaire["B2"] = statuscount["expiresoon"]
	wssommaire["B3"] = statuscount["validitytoolong"]
	wssommaire["B4"] = statuscount["notmatch"]
	wssommaire["B5"] = statuscount["wildcard"]
	wssommaire["B6"] = statuscount["ok"]
	pie = PieChart()
	labels = Reference(wssommaire, min_col=1, min_row=1, max_col=1, max_row=6)
	data = Reference(wssommaire, min_col=2, min_row=1, max_row=6, max_col=2)
	series = Series(data, title="Vue sommaire des certificats SSL")
	pie.append(series)
	pie.height = 11
	pie.width = 14
	pie.set_categories(labels)
	pie.dataLabels = label.DataLabelList()
	pie.dataLabels.showPercent = True
	pie.title = "Vue sommaire des certificats SSL"
	wssommaire.add_chart(pie, "D4")
	wb.save(pathsave+'StatusCertificates.xlsx')

def sslExpirationDate(address, port):
	if DUREE_ENTRE_CHAQUE_REQUETE != 0:
		sleep(DUREE_ENTRE_CHAQUE_REQUETE)
	try:
		conn = ssl.create_connection((address, int(port)))
		context = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
		sock = context.wrap_socket(conn, server_hostname=address)
		certificate = ssl.DER_cert_to_PEM_cert(sock.getpeercert(True))
		certLoad = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, certificate)
		certinfo = {}
		certinfo["notBefore"] = datetime.datetime.strptime(certLoad.get_notBefore().decode('ascii'), '%Y%m%d%H%M%SZ')
		certinfo["notAfter"] = datetime.datetime.strptime(certLoad.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')
		certinfo["serialNumber"] = certLoad.get_serial_number()
		if certLoad.get_issuer().CN != None:
			certinfo["deliver"] = certLoad.get_issuer().CN
		elif certLoad.get_issuer().O != None:
			certinfo["deliver"] = certLoad.get_issuer().O
		else:
			certinfo["deliver"] = "Introuvable"
		if certLoad.get_subject().CN != None:
			certinfo["deliverfor"] = certLoad.get_subject().CN
			if certLoad.get_subject().CN.lower() != address.lower():
				if certLoad.get_subject().CN.lower() == "*."+'.'.join(address.lower().split('.')[-2:]):
					certinfo["status"] = "wildcard"
				else:
					for index in range(0, certLoad.get_extension_count()):
						if "subjectAltName" in certLoad.get_extension(index).get_short_name():
							tmp = ''.join([current if ord(current) < 128 else ' ' for current in certLoad.get_extension(index).get_data().lower()])#Remove non ASCII char
							tmp = ''.join([current for current in tmp if current in string.printable])#Remove non printable char
							if "*."+'.'.join(address.lower().split('.')[-2:]) in ' '.join(tmp.split()).split(' '):
								certinfo["status"] = "wildcard"
							elif address.lower() not in ' '.join(tmp.split()).split(' '):
								return "notmatch"
							break
		else:
			certinfo["deliverfor"] = "Introuvable"
		return certinfo
	except socket.error as err:
		if ("WRONG_SSL_VERSION" in str(err) or "UNSUPPORTED_PROTOCOL" in str(err)):
			try:
				conn = ssl.create_connection((address, int(port)))
				context = ssl.SSLContext(ssl.PROTOCOL_TLSv1)
				sock = context.wrap_socket(conn, server_hostname=address)
				certificate = ssl.DER_cert_to_PEM_cert(sock.getpeercert(True))
				certLoad = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, certificate)
				certinfo = {}
				certinfo["notBefore"] = datetime.datetime.strptime(certLoad.get_notBefore().decode('ascii'), '%Y%m%d%H%M%SZ')
				certinfo["notAfter"] = datetime.datetime.strptime(certLoad.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')
				certinfo["serialNumber"] = certLoad.get_serial_number()
				if certLoad.get_issuer().CN != None:
					certinfo["deliver"] = certLoad.get_issuer().CN
				elif certLoad.get_issuer().O != None:
					certinfo["deliver"] = certLoad.get_issuer().O
				else:
					certinfo["deliver"] = "Introuvable"
				if certLoad.get_subject().CN != None:
					certinfo["deliverfor"] = certLoad.get_subject().CN
					if certLoad.get_subject().CN.lower() != address.lower():
						for index in range(0, certLoad.get_extension_count()):
							if certLoad.get_subject().CN.lower() == "*."+'.'.join(address.lower().split('.')[-2:]):
								certinfo["status"] = "wildcard"
							else:
								for index in range(0, certLoad.get_extension_count()):
									if "subjectAltName" in certLoad.get_extension(index).get_short_name():
										tmp = ''.join([current if ord(current) < 128 else ' ' for current in certLoad.get_extension(index).get_data().lower()])#Remove non ASCII char
										tmp = ''.join([current for current in tmp if current in string.printable])#Remove non printable char
										if "*."+'.'.join(address.lower().split('.')[-2:]) in ' '.join(tmp.split()).split(' '):
											certinfo["status"] = "wildcard"
										elif address.lower() not in ' '.join(tmp.split()).split(' '):
											return "notmatch"
										break
				else:
					certinfo["deliverfor"] = "Introuvable"
				return certinfo
			except ssl.SSLError as err:
				return "error"
			except socket.error as err:
				return "error"
			except:
				print "error"
				return "error"
		elif ("[Errno 0]" in str(err)):
			try:
				conn = ssl.create_connection((address, int(port)))
				context = ssl.SSLContext(ssl.PROTOCOL_TLSv1_1)
				sock = context.wrap_socket(conn, server_hostname=address)
				certificate = ssl.DER_cert_to_PEM_cert(sock.getpeercert(True))
				certLoad = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, certificate)
				certinfo = {}
				certinfo["notBefore"] = datetime.datetime.strptime(certLoad.get_notBefore().decode('ascii'), '%Y%m%d%H%M%SZ')
				certinfo["notAfter"] = datetime.datetime.strptime(certLoad.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')
				certinfo["serialNumber"] = certLoad.get_serial_number()
				if certLoad.get_issuer().CN != None:
					certinfo["deliver"] = certLoad.get_issuer().CN
				elif certLoad.get_issuer().O != None:
					certinfo["deliver"] = certLoad.get_issuer().O
				else:
					certinfo["deliver"] = "Introuvable"
				if certLoad.get_subject().CN != None:
					certinfo["deliverfor"] = certLoad.get_subject().CN
					if certLoad.get_subject().CN.lower() != address.lower():
						if certLoad.get_subject().CN.lower() == "*."+'.'.join(address.lower().split('.')[-2:]):
							certinfo["status"] = "wildcard"
						else:
							for index in range(0, certLoad.get_extension_count()):
								if "subjectAltName" in certLoad.get_extension(index).get_short_name():
									tmp = ''.join([current if ord(current) < 128 else ' ' for current in certLoad.get_extension(index).get_data().lower()])#Remove non ASCII char
									tmp = ''.join([current for current in tmp if current in string.printable])#Remove non printable char
									if "*."+'.'.join(address.lower().split('.')[-2:]) in ' '.join(tmp.split()).split(' '):
										certinfo["status"] = "wildcard"
									elif address.lower() not in ' '.join(tmp.split()).split(' '):
										return "notmatch"
									break
				else:
					certinfo["deliverfor"] = "Introuvable"
				return certinfo
			except ssl.CertificateError as err:
				if DEBUG:
					print str(err)
				if ("doesn't match either of" in str(err)):
					return "notmatch"
				return "error"
			except socket.error as err:
				if ("[Errno 0]" in str(err)):
					try:
						conn = ssl.create_connection((address, int(port)))
						context = ssl.SSLContext(ssl.PROTOCOL_TLSv1)
						sock = context.wrap_socket(conn, server_hostname=address)
						certificate = ssl.DER_cert_to_PEM_cert(sock.getpeercert(True))
						certLoad = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, certificate)
						certinfo = {}
						certinfo["notBefore"] = datetime.datetime.strptime(certLoad.get_notBefore().decode('ascii'), '%Y%m%d%H%M%SZ')
						certinfo["notAfter"] = datetime.datetime.strptime(certLoad.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')
						certinfo["serialNumber"] = certLoad.get_serial_number()
						if certLoad.get_issuer().CN != None:
							certinfo["deliver"] = certLoad.get_issuer().CN
						elif certLoad.get_issuer().O != None:
							certinfo["deliver"] = certLoad.get_issuer().O
						else:
							certinfo["deliver"] = "Introuvable"
						if certLoad.get_subject().CN.lower() != address.lower():
							if certLoad.get_subject().CN.lower() == "*."+'.'.join(address.lower().split('.')[-2:]):
								certinfo["status"] = "wildcard"
							else:
								for index in range(0, certLoad.get_extension_count()):
									if "subjectAltName" in certLoad.get_extension(index).get_short_name():
										tmp = ''.join([current if ord(current) < 128 else ' ' for current in certLoad.get_extension(index).get_data().lower()])#Remove non ASCII char
										tmp = ''.join([current for current in tmp if current in string.printable])#Remove non printable char
										if "*."+'.'.join(address.lower().split('.')[-2:]) in ' '.join(tmp.split()).split(' '):
											certinfo["status"] = "wildcard"
										elif address.lower() not in ' '.join(tmp.split()).split(' '):
											return "notmatch"
										break
						return certinfo
					except socket.error as err:
						if DEBUG:
							print str(err)
						if ("[Errno 0]" in str(err)):
							return "INVALID CERT"
					except:
						print "error"
						return "error"
				print str(err)+"second socket"
				return "error"
			except:
				print "error"
				return "error"
		elif "Name or service not known" in str(err):
			return "errresolution"
		elif "Connection refused" in str(err):
			return "CONN REFUSED"
		if DEBUG:
			print str(err)
		return "error"
	except ssl.CertificateError as err:
		if ("doesn't match either of" in str(err)):
			return "notmatch"
		if DEBUG:
			print str(err)
		return "error"
	except ssl.gaierror as err:
		if "Name or service not known" in str(err):
			return "errresolution"
	except:
		print "error"
		return "error"

def serialToHex(serialnumber):
	tmp = format(serialnumber, 'x')
	if len(str(tmp)) % 2 != 0:
		tmp = "0"+tmp
	final = ':'.join(tmp[current:current+2] for current in range(0, len(tmp), 2))
	return final.upper()

def sslExpirationCalcul(address, port):
	expirationdate = sslExpirationDate(address, port)
	if type(expirationdate) != str:
		certcalcul = {}
		if "status" in expirationdate:
			certcalcul["status"] = expirationdate["status"]
		certcalcul["notAfter"] = expirationdate["notAfter"]
		certcalcul["notBefore"] = expirationdate["notBefore"]
		certcalcul["serialNumber"] = serialToHex(expirationdate["serialNumber"])
		certcalcul["deliver"] = expirationdate["deliver"]
		certcalcul["deliverfor"] = expirationdate["deliverfor"]
		certcalcul["deltaToday"] = expirationdate["notAfter"] - datetime.datetime.utcnow()
		certcalcul["deltaValidity"] = expirationdate["notAfter"] - expirationdate["notBefore"]
		return certcalcul
	else:
		return expirationdate

def sslExpirationStatus(address, port, days_check, validity_period, return_value):
	getdate = sslExpirationCalcul(address, port)
	result = {}
	result["domain"] = address
	result["port"] = port
	if type(getdate) == str:
		result["status"] = getdate
		result["periodevalidity"] = "error"
		result["deltaToday"] = "error"
		result["deliver"] = "error"
		result["notAfter"] = "error"
		result["notBefore"] = "error"
		result["serialNumber"] = "error"
		result["deliverfor"] = "error"
		return_value[0] = result
	else:
		result["notAfter"] = getdate["notAfter"]
		result["notBefore"] = getdate["notBefore"]
		result["serialNumber"] = getdate["serialNumber"]
		result["deliver"] = getdate["deliver"]
		result["deltaToday"] = getdate["deltaToday"].days
		result["periodevalidity"] = getdate["deltaValidity"].days
		result["deliverfor"] = getdate["deliverfor"]
		if getdate["deltaToday"] <= datetime.timedelta(days=0):
			result["status"] = "expire"
		elif getdate["deltaToday"] < datetime.timedelta(days=days_check):
			result["status"] = "expiresoon"
		else:
			if getdate["deltaValidity"] > datetime.timedelta(days=validity_period):
				result["status"] = "validitytoolong"
			else:
				if "status" in getdate:
					result["status"] = getdate["status"]
				else:
					result["status"] = "ok"
	return_value[0] = result

#Permet de vérifier l association port+serial
def checkportserial(listresult, currentcheck):
	for current in listresult:
		if current["serialNumber"] == currentcheck["serialNumber"]:
			if currentcheck["domain"] not in current["domain"].split('~'):
				current["domain"] = current["domain"]+'~'+currentcheck["domain"]
			if currentcheck["port"] not in current["port"].split('~'):
				current["port"] = current["port"]+'~'+currentcheck["port"]
			return listresult
	listresult.append(currentcheck)
	return listresult

if len(sys.argv) < 2:
	print "usage: python jajertificats.py PATH_FICHIER_LISTE_DOMAINE"
	sys.exit(1)
allresult = []
countcheck = 1
countlistcheck = len(open(sys.argv[1]).readlines(  ))*len(LIST_PORT)
with open(sys.argv[1], "r") as f:
    for currentdomain in f:
		currentdomain = currentdomain.replace('\n', '')
		for currentport in LIST_PORT:
			domaininprogress = currentdomain
			manager = multiprocessing.Manager()
			resultthread = manager.dict()
			p = multiprocessing.Process(target=sslExpirationStatus, args=(domaininprogress.encode('utf8'), currentport, EXPIRATION_DAYS, PERIODE_DE_VALIDITE, resultthread))
			p.start()
			p.join(TIMEOUT_REQUEST)
			if not p.is_alive():
				result = resultthread[0]
				if result["status"] == "errresolution":
					errresolutiondomain.append(domaininprogress)
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[ERR RESOLUTION]\033[1;m'+"       Impossible de joindre "+domaininprogress+u", les autres ports ne seront pas vérifiés"
				elif result["status"] == "CONN REFUSED":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[CONN REFUSEE]\033[1;m'+u"         Connexion refusée pour "+domaininprogress+" sur le port "+currentport
				elif result["status"] == "INVALID CERT":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[CERT INVALIDE]\033[1;m'+u"        Certificat invalide pour "+domaininprogress+" sur le port "+currentport
				elif result["status"] == "error":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[ERREUR]\033[1;m'+"               Impossible de recuperer le certificat pour "+domaininprogress+" sur le port "+currentport
				elif result["status"] == "notmatch":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[HOSTNAME INVALIDE]\033[1;m'+"    Le nom de l\'hostname ne correspond pas au certificat pour "+domaininprogress+" sur le port "+currentport
				elif result["status"] == "expire":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[EXPIRE]\033[1;m'+"               Certificat pour "+domaininprogress+u" expiré depuis "+str(result["deltaToday"]*-1)+" jours sur le port "+currentport
				elif result["status"] == "expiresoon":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;33m[EXPIRE BIENTOT]\033[1;m'+"       Certificat pour "+domaininprogress+" expire dans "+str(result["deltaToday"])+" jours sur le port "+currentport
				elif result["status"] == "validitytoolong":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;33m[VALIDITE TROP LONGUE]\033[1;m'+" Certificat pour "+domaininprogress+u" a une période de validité de "+str(result["periodevalidity"])+" jours sur le port "+currentport
				elif result["status"] == "wildcard":
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;34m[WILDCARD]\033[1;m'+"             Certificat pour "+domaininprogress+" OK (wildcard) sur le port "+currentport
				else:
					print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;34m[OK]\033[1;m'+"                   Certificat pour "+domaininprogress+" OK sur le port "+currentport
			else:
				try:
					p.terminate()
				except Exception as e:
					print e
				result = {}
				result["domain"] = domaininprogress
				result["status"] = "timeout"
				result["port"] = currentport
				result["deliver"] = "error"
				result["deliverfor"] = "error"
				result["deltaToday"] = "error"
				result["notAfter"] = "error"
				result["notBefore"] = "error"
				result["serialNumber"] = "error"
				result["periodevalidity"] = "error"
				print "[Check "+str(countcheck).zfill(len(str(countlistcheck)))+" / "+str(countlistcheck)+"] "+'\033[1;31m[TIMEOUT]\033[1;m'+"              Timeout pour "+domaininprogress+" sur le port "+currentport
			if result["serialNumber"] != "error":
				allresult = checkportserial(allresult, result)
			else:
				allresult.append(result)
			countcheck += 1

print "\nSauvegarde des resultats en cours..."
writeresult(allresult, './')
print "Resultats disponible dans le fichier StatusCertificates.xlsx"
